"""文档加载器 — 支持 txt/pdf/doc/docx/xlsx/xls 多格式"""

from pathlib import Path

from loguru import logger


async def load_document(file_path: str | Path, file_type: str | None = None) -> str:
    """加载文档内容，自动识别格式

    Args:
        file_path: 文件路径
        file_type: 文件扩展名（如 txt/pdf/docx），为 None 时自动从路径推断

    Returns:
        文档纯文本内容
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = (file_type or path.suffix.lstrip(".")).lower()

    loaders = {
        "txt": _load_txt,
        "pdf": _load_pdf,
        "docx": _load_docx,
        "doc": _load_doc,
        "xlsx": _load_xlsx,
        "xls": _load_xlsx,
    }

    loader = loaders.get(ext)
    if loader is None:
        raise ValueError(f"不支持的文件格式: {ext}，支持: {', '.join(loaders.keys())}")

    content = await loader(path)
    logger.debug(f"文档加载完成: {path.name}, 长度: {len(content)}")
    return content


async def _load_txt(path: Path) -> str:
    """加载 txt 文件（UTF-8 优先，失败回退系统编码）"""
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        logger.warning(f"UTF-8 读取失败，尝试系统编码: {path}")
        return path.read_text(encoding="gbk")


async def _load_pdf(path: Path) -> str:
    """加载 PDF 文件"""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


async def _load_docx(path: Path) -> str:
    """加载 docx 文件"""
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


async def _load_doc(path: Path) -> str:
    """加载 doc 文件（尝试 docx 方式，失败则记录警告）"""
    try:
        return await _load_docx(path)
    except Exception:
        logger.warning(f".doc 格式支持有限，请转换为 .docx 或 .txt: {path}")
        # 尝试按纯文本读取
        return await _load_txt(path)


async def _load_xlsx(path: Path) -> str:
    """加载 xlsx/xls 文件"""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                rows.append(row_text)
    wb.close()
    return "\n".join(rows)
